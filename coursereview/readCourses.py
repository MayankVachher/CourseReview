import os
import re
import json
from openpyxl import load_workbook


courses = {}

basepath = os.path.dirname(__file__)
filepath = os.path.abspath(os.path.join(basepath, "..", "static", "data", "courseList.xlsx"))
jsonfilepath = os.path.abspath(os.path.join(basepath, "..", "static", "data", "courseData.json"))

wb = load_workbook(filepath)
ws = wb.get_sheet_by_name(name='Course numbers')

pattern = re.compile('^[A-Z]{3}[0-9]{3}$')

for rownum in range(1, ws.get_highest_row()):
    cell_data = str(ws.cell(row=rownum, column=1).value)
    if cell_data == "None":
        continue
    elif re.match(pattern, cell_data) is not None:
        courseArea = cell_data[:-3]
        courseID = cell_data[3:]
        courseName = str(ws.cell(row=rownum, column=2).value)
        if not courses.has_key(courseArea):
            courses[courseArea] = []
        courses[courseArea].append({
            'ID': courseID,
            'Name': courseName})

with open(jsonfilepath, 'w') as outfile:
    json.dump(courses, outfile)

print "\nCourses successfully written to ", jsonfilepath, "\n"